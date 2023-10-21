import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# Function to read data from Excel file
def read_data(file_path, sheet_name='sheet_name'):
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
        return df
    except FileNotFoundError:
        print("Error: Input file not found.")
        return None

# Load data using the read_data function
input_file = 'General.xlsx'
df = read_data(input_file, sheet_name='GENERAL')

output_file = 'Output-General.xlsx'
# Convert necessary columns to numeric, coerce errors to NaN
numeric_columns = ['InvWgt', 'Shares', 'Price']
df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')

# Drop rows with NaN values in numeric columns
df = df.dropna(subset=numeric_columns)

# Calculate Market Cap
df['MARKET CAP'] = df['Shares'] * df['Price'] * df['InvWgt'] / 100
# Calculate UnAdjusted Market Cap
df['UnAdjustedMarketCap'] = df['Shares'] * df['Price'] * df['InvWgt'] / 100

# Calculate Total Market Cap
total_market_cap = df['MARKET CAP'].sum()

# Calculate Initial Weights
df['Weight'] = (df['MARKET CAP'] / total_market_cap) * 100
df['UnCapWeight'] = (df['MARKET CAP'] / total_market_cap) * 100

# Capping limit
CAP_LIMIT = 25

# Maximum iterations to prevent infinite loop
MAX_ITERATIONS = 1000

# Iteratively adjust market cap and weights
for i in range(MAX_ITERATIONS):
    excess_weight = df['Weight'].max() - CAP_LIMIT
    if excess_weight <= 0:
        break

    total_excess_weight = df[df['Weight'] > CAP_LIMIT]['Weight'].sum()
    redistribution_prop = excess_weight / total_excess_weight

    # Update market cap for constituents with excess weight
    df.loc[df['Weight'] > CAP_LIMIT, 'MARKET CAP'] *= (1 - redistribution_prop)

    # Recalculate weights based on adjusted market cap
    total_market_cap = df['MARKET CAP'].sum()
    df['Weight'] = (df['MARKET CAP'] / total_market_cap) * 100
    df['CappWeight'] = df['Weight']

# Drop the 'Weight' column from the DataFrame
df.drop('Weight', axis=1, inplace=True)
# Rename the 'MARKET CAP' column to 'Adj Capped MarkCAP'
df.rename(columns={'MARKET CAP': 'AdjustedMarketCap'}, inplace=True)
##############################################################################################
# Calculate CappingFactor (Ci)
print(df[['AdjustedMarketCap', 'UnAdjustedMarketCap']])
df['CappingFactor'] = df['AdjustedMarketCap'] / df['UnAdjustedMarketCap']
# Calculate CappingFactor (Ci) with 8 decimal points
df['CappingFactor'] = ((df['AdjustedMarketCap'] / df['UnAdjustedMarketCap']*100).round(4))

# Define the desired column order
desired_columns = ['Code', 'Company Name', 'InvWgt', 'Shares', 'Price', 'UnAdjustedMarketCap', 'AdjustedMarketCap', 'UnCapWeight', 'CappWeight', 'CappingFactor']
# Reorder the columns in the DataFrame
df = df[desired_columns]

# Create a Pandas Excel writer using XlsxWriter as the engine
excel_writer = pd.ExcelWriter(output_file, engine='openpyxl')
excel_writer.book = Workbook()

# Convert DataFrame to worksheet using dataframe_to_rows
rows = dataframe_to_rows(df, index=False, header=True)
worksheet = excel_writer.book.active
# Set freeze panes to freeze the first row and the first two columns
worksheet.freeze_panes = 'C2'

# Apply number formats to specific columns
number_format = {
    'Price': '#,##0.0000',         # Format Price column with 4 decimal places
    'CappingFactor': '#,##0.0000',  # Format CappingFactor column with 4 decimal places
    'UnCapWeight': '#,##0.00',      # Format UnCapWeight column with 2 decimal places
    'CappWeight': '#,##0.00'        # Format CappWeight column with 2 decimal places
}

# Convert 'UnCapWeight' and 'CappWeight' columns to numeric
df['UnCapWeight'] = pd.to_numeric(df['UnCapWeight'], errors='coerce')
df['CappWeight'] = pd.to_numeric(df['CappWeight'], errors='coerce')
df['Shares'] = pd.to_numeric(df['Shares'], errors='coerce')
# Format 'Shares' column with thousand separator and without decimal points
#df['Shares'] = df['Shares'].apply(lambda x: '{:,.0f}'.format(x))

# Round the 'Price' column to 4 decimal places
df['Price'] = df['Price'].round(4)

# Round the numeric columns to the desired decimal places
df['UnCapWeight'] = df['UnCapWeight'].round(2)
df['CappWeight'] = df['CappWeight'].round(2)

# Write data to worksheet
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        worksheet.cell(row=r_idx, column=c_idx, value=value)

# Autofit columns
for column_cells in worksheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

# Add table style to the table
table = Table(displayName="Table1", ref=worksheet.dimensions)
style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
table.tableStyleInfo = style
worksheet.add_table(table)

# Save the Excel file
excel_writer.save()
print("Data exported to", output_file)
